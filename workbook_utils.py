from __future__ import annotations

import io
import math
import re
import xml.etree.ElementTree as ET
from collections import OrderedDict
from datetime import date, datetime
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.datetime import to_excel

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XML_SPACE = "http://www.w3.org/XML/1998/namespace"


# ---------------------------------------------------------------------------
# Canonical workbook fields
# ---------------------------------------------------------------------------
CORE_FIELDS = [
    "Deal Number",
    "Deal Name",
    "Location",
    "Bridge / Term",
    "Type",
    "Segment",
    "Financing",
    "Current UPB",
    "Salesforce Valuation Date",
    "Salesforce As-Is Valuation",
    "Salesforce Implied As-Is LTV",
    "Loan Commitment",
    "Salesforce ARV",
    "Salesforce ARV LTV",
    "12/31 Px*UPB",
    "12/31 Px",
    "Remaining Commitment",
    "Number Of Units",
    "Asset Manager",
    "Maturity Date",
    "Next Payment Date",
    "12/31 NPL",
    "3/31 NPL",
    "6/30 NPL",
    "Current DQ Status",
    "MBA Loan List Convention",
]

EXPORTABLE_FIELDS = [
    "Resolution Likelihood",
    "Estimated Resolution Timing",
    "Estimated Liquidation Timing",
    "Expected Resolution Type",
    "Resolution for 6/30 Performance",
    "Addt'l NPL Comment",
    "Mod Needed Prior to Final Resolution (Y/N)",
    "FCL Needed Prior to Final Resolution (Y/N)",
    "FCL Date",
    "Expected Final Resolution",
    "Resolution Timing",
    "Liquidity Event Timing",
    "Pref Deal (Y/N)",
    "Pref Amount ($)",
    "Liquidity Event ($)",
    "CV Inspected in Last 6 months? (Y/N)",
    "New Valuation in 1Q26 (Y/N)",
    "Valuation Type",
    "As Is Appraised Value",
    "As Stabilized Value",
    "Report Date",
    "Date of Valuation",
    "Is Appraisal Finalized? (Y/N)",
    "Third Party Offer Amount",
    "Third Party Offer Note",
    "Comments",
]

READONLY_CONTEXT_FIELDS = [
    "Current Salesforce AM Comment",
]

LONG_TEXT_FIELDS = {
    "Addt'l NPL Comment",
    "Third Party Offer Note",
    "Current Salesforce AM Comment",
    "Comments",
    "Previous Weekly Comment",
    "This Week Comment",
    "Comments Preview",
}

DATE_OR_TEXT_FIELDS = {
    "FCL Date",
    "Report Date",
    "Date of Valuation",
}

NUMERIC_OR_TEXT_FIELDS = {
    "Pref Amount ($)",
    "Liquidity Event ($)",
    "As Is Appraised Value",
    "As Stabilized Value",
    "Third Party Offer Amount",
}

COMMENT_STORAGE_FIELD = "Addt'l NPL Comment"
MANUAL_ENTRY_PICKLIST_FIELDS = {"Expected Resolution Type", "Valuation Type"}

GRID_CONTEXT_HEADERS = [
    "Deal Number",
    "Deal Name",
    "Asset Manager",
    "Location",
    "Current DQ Status",
    "6/30 NPL",
    "Current UPB",
]

SECTION_MAP = OrderedDict(
    {
        "Resolution / projection": [
            "Resolution Likelihood",
            "Estimated Resolution Timing",
            "Estimated Liquidation Timing",
            "Expected Resolution Type",
            "Resolution for 6/30 Performance",
            "Addt'l NPL Comment",
        ],
        "Resolution path": [
            "Mod Needed Prior to Final Resolution (Y/N)",
            "FCL Needed Prior to Final Resolution (Y/N)",
            "FCL Date",
            "Expected Final Resolution",
            "Resolution Timing",
            "Liquidity Event Timing",
            "Pref Deal (Y/N)",
            "Pref Amount ($)",
            "Liquidity Event ($)",
        ],
        "Valuation": [
            "CV Inspected in Last 6 months? (Y/N)",
            "New Valuation in 1Q26 (Y/N)",
            "Valuation Type",
            "As Is Appraised Value",
            "As Stabilized Value",
            "Report Date",
            "Date of Valuation",
            "Is Appraisal Finalized? (Y/N)",
            "Third Party Offer Amount",
            "Third Party Offer Note",
        ],
    }
)

MEETING_STATUS_OPTIONS = ["Open", "Discussed", "Parking Lot", "Done"]
COMMENT_TEMPLATE_OPTIONS = [
    "",
    "No change from last week",
    "Negotiating payoff / DPO",
    "Foreclosure / receiver path advancing",
    "LOI / bid in hand",
    "Modification / extension in process",
    "Valuation ordered / received",
    "Resolved / no longer NPL",
    "Counsel / title / documentation follow-up",
]

FIELD_ALIASES = {
    "Deal Number": "Deal Number",
    "Deal Name": "Deal Name",
    "Location": "Location",
    "Bridge / Term": "Bridge / Term",
    "Type": "Type",
    "Segment": "Segment",
    "Financing": "Financing",
    "Salesforce Valuation Date": "Salesforce Valuation Date",
    "Salesforce As-Is Valuation": "Salesforce As-Is Valuation",
    "Salesforce Implied As-Is LTV": "Salesforce Implied As-Is LTV",
    "Loan Commitment": "Loan Commitment",
    "Salesforce ARV": "Salesforce ARV",
    "Salesforce ARV LTV": "Salesforce ARV LTV",
    "12/31 Px*UPB": "12/31 Px*UPB",
    "12/31 Px": "12/31 Px",
    "Remaining Commitment": "Remaining Commitment",
    "Number Of Units": "Number Of Units",
    "Asset Manager": "Asset Manager",
    "Maturity Date": "Maturity Date",
    "Next Payment Date": "Next Payment Date",
    "12/31 NPL": "12/31 NPL",
    "3/31 NPL": "3/31 NPL",
    "6/30 NPL": "6/30 NPL",
    "MBA Loan List Convention": "MBA Loan List Convention",
    "Resolution Likelihood": "Resolution Likelihood",
    "Resolution for 6/30 Performance": "Resolution for 6/30 Performance",
    "Estimated Resolution Timing": "Estimated Resolution Timing",
    "Estimated Liquidation Timing": "Estimated Liquidation Timing",
    "Expected Resolution Type": "Expected Resolution Type",
    "Addt'l NPL Comment": "Addt'l NPL Comment",
    "Mod Needed Prior to Final Resolution (Y/N)": "Mod Needed Prior to Final Resolution (Y/N)",
    "FCL Needed Prior to Final Resolution (Y/N)": "FCL Needed Prior to Final Resolution (Y/N)",
    "FCL Date": "FCL Date",
    "Expected Final Resolution": "Expected Final Resolution",
    "Resolution Timing": "Resolution Timing",
    "Liquidity Event Timing": "Liquidity Event Timing",
    "Pref Deal (Y/N)": "Pref Deal (Y/N)",
    "Pref Amount ($)": "Pref Amount ($)",
    "Liquidity Event ($)": "Liquidity Event ($)",
    "CV Inspected in Last 6 months? (Y/N)": "CV Inspected in Last 6 months? (Y/N)",
    "New Valuation in 1Q26 (Y/N)": "New Valuation in 1Q26 (Y/N)",
    "Valuation Type": "Valuation Type",
    "As Is Appraised Value": "As Is Appraised Value",
    "As Stabilized Value": "As Stabilized Value",
    "Report Date": "Report Date",
    "Date of Valuation": "Date of Valuation",
    "Is Appraisal Finalized? (Y/N)": "Is Appraisal Finalized? (Y/N)",
    "Third Party Offer Amount": "Third Party Offer Amount",
    "Third Party Offer Note": "Third Party Offer Note",
    "Comments": "Comments",
    # Property sheet aliases
    "Servicer ID": "Servicer ID",
    "Servicer": "Servicer",
    "SF Yardi ID": "SF Yardi ID",
    "Asset ID": "Asset ID",
    "Address": "Address",
    "City": "City",
    "State": "State",
    "# Units": "# Units",
    "First Funded Date": "First Funded Date",
    "REO Date": "REO Date",
    "Days Past Due": "Days Past Due",
    "Borrower Name": "Borrower Name",
    "Account": "Account",
}


def sanitize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def canonicalize_header(value: Any) -> str:
    text = sanitize_header(value)
    if not text:
        return ""
    if re.fullmatch(r"\d{1,2}/\d{1,2} UPB", text):
        return "Current UPB"
    if re.fullmatch(r"\d{1,2}/\d{1,2} DQ Status", text):
        return "Current DQ Status"
    if re.fullmatch(r"\d{1,2}/\d{1,2} Salesforce AM Comment", text):
        return "Current Salesforce AM Comment"
    return FIELD_ALIASES.get(text, text)


# ---------------------------------------------------------------------------
# Workbook detection / parsing
# ---------------------------------------------------------------------------
def get_sheet_info(file_bytes: bytes) -> dict[str, dict[str, str]]:
    with ZipFile(io.BytesIO(file_bytes)) as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

    rel_map = {
        rel.get("Id"): rel.get("Target")
        for rel in rels_root.findall(f"{{{REL_NS}}}Relationship")
    }

    sheet_info: dict[str, dict[str, str]] = {}
    for sheet in workbook_root.find(f"{{{MAIN_NS}}}sheets"):
        name = sheet.get("name", "")
        rid = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        target = rel_map.get(rid, "")
        if target and not target.startswith("xl/"):
            target = f"xl/{target}"
        sheet_info[name] = {
            "state": sheet.get("state", "visible"),
            "rid": rid,
            "path": target,
        }
    return sheet_info


def get_hidden_rows(file_bytes: bytes, sheet_path: str) -> set[int]:
    with ZipFile(io.BytesIO(file_bytes)) as zf:
        root = ET.fromstring(zf.read(sheet_path))
    hidden_rows: set[int] = set()
    sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
    if sheet_data is None:
        return hidden_rows
    for row in sheet_data.findall(f"{{{MAIN_NS}}}row"):
        if row.get("hidden") == "1":
            try:
                hidden_rows.add(int(row.get("r")))
            except Exception:
                continue
    return hidden_rows


def detect_sheet_names(file_bytes: bytes) -> dict[str, str | None]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_info = get_sheet_info(file_bytes)

    deal_candidates: list[tuple[int, str]] = []
    property_candidates: list[tuple[int, str]] = []

    for ws in wb.worksheets:
        state = sheet_info.get(ws.title, {}).get("state", "visible")
        visible_bonus = 100 if state == "visible" else 0

        try:
            row2 = [canonicalize_header(v) for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
        except StopIteration:
            row2 = []
        try:
            row4 = [canonicalize_header(v) for v in next(ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        except StopIteration:
            row4 = []

        row2_set = set(row2)
        row4_set = set(row4)

        if {"Deal Number", "Deal Name", "Asset Manager"}.issubset(row2_set):
            score = visible_bonus
            if "Loan List" in ws.title:
                score += 10
            deal_candidates.append((score, ws.title))

        if {"Deal Number", "Servicer ID", "Address"}.issubset(row4_set):
            score = visible_bonus
            if "Property List" in ws.title:
                score += 10
            property_candidates.append((score, ws.title))

    return {
        "deal_sheet": max(deal_candidates)[1] if deal_candidates else None,
        "property_sheet": max(property_candidates)[1] if property_candidates else None,
    }


def detect_snapshot_label(raw_headers: list[str]) -> str | None:
    for header in raw_headers:
        if re.fullmatch(r"\d{1,2}/\d{1,2} UPB", header):
            return header.replace(" UPB", "")
    for header in raw_headers:
        if re.fullmatch(r"\d{1,2}/\d{1,2} DQ Status", header):
            return header.replace(" DQ Status", "")
    return None


def normalize_number(value: Any) -> float | None:
    if value in (None, "", "N/A", "NA", "-"):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def summarize_properties(properties_df: pd.DataFrame) -> pd.DataFrame:
    if properties_df.empty or "Deal Number" not in properties_df.columns:
        return pd.DataFrame()

    working = properties_df.copy()
    for col in ["# Units", "Current UPB", "Salesforce As-Is Valuation", "Salesforce ARV", "Days Past Due"]:
        if col in working.columns:
            working[col] = working[col].apply(normalize_number)

    def join_unique(series: pd.Series, limit: int = 4) -> str:
        values = []
        for item in series.dropna().astype(str):
            item = item.strip()
            if item and item not in values:
                values.append(item)
        if len(values) <= limit:
            return ", ".join(values)
        return ", ".join(values[:limit]) + f" +{len(values) - limit} more"

    summary = (
        working.groupby("Deal Number", dropna=False)
        .agg(
            **{
                "Property Count": ("Deal Number", "size"),
                "Property Units": ("# Units", "sum"),
                "Property UPB": ("Current UPB", "sum"),
                "Property As-Is Value": ("Salesforce As-Is Valuation", "sum"),
                "Property ARV": ("Salesforce ARV", "sum"),
                "Max Days Past Due": ("Days Past Due", "max"),
                "Property Cities": ("City", join_unique),
            }
        )
        .reset_index()
    )
    return summary


def parse_deal_workbook(file_bytes: bytes) -> dict[str, Any]:
    sheet_names = detect_sheet_names(file_bytes)
    deal_sheet = sheet_names["deal_sheet"]
    property_sheet = sheet_names["property_sheet"]
    if not deal_sheet:
        raise ValueError("Could not find the main deal sheet in the uploaded workbook.")

    wb_values = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    wb_formula = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    sheet_info = get_sheet_info(file_bytes)

    deals_values_ws = wb_values[deal_sheet]
    deals_formula_ws = wb_formula[deal_sheet]
    deal_sheet_path = sheet_info[deal_sheet]["path"]
    hidden_rows = get_hidden_rows(file_bytes, deal_sheet_path)

    raw_headers = [sanitize_header(v) for v in next(deals_values_ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    canonical_headers = [canonicalize_header(v) for v in raw_headers]
    snapshot_label = detect_snapshot_label(raw_headers)
    raw_header_by_canonical = {
        canonical_headers[idx]: raw_headers[idx]
        for idx in range(len(canonical_headers))
        if canonical_headers[idx]
    }
    col_letters = {
        canonical_headers[idx]: get_column_letter(idx + 1)
        for idx in range(len(canonical_headers))
        if canonical_headers[idx]
    }

    records: list[dict[str, Any]] = []
    original_editor_values: dict[int, dict[str, str]] = {}
    cell_meta: dict[tuple[int, str], dict[str, Any]] = {}

    value_rows = deals_values_ws.iter_rows(min_row=3, max_row=deals_values_ws.max_row, values_only=False)
    formula_rows = deals_formula_ws.iter_rows(min_row=3, max_row=deals_formula_ws.max_row, values_only=False)

    for value_row, formula_row in zip(value_rows, formula_rows):
        excel_row = value_row[0].row
        if value_row[0].value in (None, ""):
            continue

        record: dict[str, Any] = {
            "_excel_row": excel_row,
            "_workbook_hidden": excel_row in hidden_rows,
        }
        editor_values_for_row: dict[str, str] = {}

        for idx, (value_cell, formula_cell) in enumerate(zip(value_row, formula_row)):
            if idx >= len(canonical_headers):
                continue
            header = canonical_headers[idx]
            if not header:
                continue
            value = value_cell.value
            record[header] = value
            if header in EXPORTABLE_FIELDS or header in READONLY_CONTEXT_FIELDS:
                editor_values_for_row[header] = editable_text(value)
                cell_meta[(excel_row, header)] = {
                    "coord": formula_cell.coordinate,
                    "data_type": formula_cell.data_type,
                    "number_format": formula_cell.number_format,
                    "original_formula": formula_cell.value if formula_cell.data_type == "f" else None,
                    "raw_value": formula_cell.value,
                    "column_letter": col_letters[header],
                }

        records.append(record)
        original_editor_values[excel_row] = editor_values_for_row

    deals_df = pd.DataFrame(records)
    if deals_df.empty:
        raise ValueError("The main deal sheet was found, but no deal rows could be read.")

    properties_df = pd.DataFrame()
    property_summary = pd.DataFrame()
    if property_sheet:
        properties_ws = wb_values[property_sheet]
        raw_prop_headers = [sanitize_header(v) for v in next(properties_ws.iter_rows(min_row=4, max_row=4, values_only=True))]
        prop_headers = [canonicalize_header(v) for v in raw_prop_headers]
        property_records = []
        for row in properties_ws.iter_rows(min_row=5, max_row=properties_ws.max_row, values_only=True):
            if row[0] in (None, ""):
                continue
            property_records.append(
                {
                    prop_headers[i]: row[i]
                    for i in range(min(len(prop_headers), len(row)))
                    if prop_headers[i]
                }
            )
        if property_records:
            properties_df = pd.DataFrame(property_records)
            property_summary = summarize_properties(properties_df)

    available_fields = [field for field in CORE_FIELDS + EXPORTABLE_FIELDS + READONLY_CONTEXT_FIELDS if field in deals_df.columns]

    return {
        "deal_sheet": deal_sheet,
        "property_sheet": property_sheet,
        "deal_sheet_path": deal_sheet_path,
        "deals_df": deals_df,
        "properties_df": properties_df,
        "property_summary": property_summary,
        "cell_meta": cell_meta,
        "original_editor_values": original_editor_values,
        "sheet_info": sheet_info,
        "canonical_headers": canonical_headers,
        "raw_headers": raw_headers,
        "raw_header_by_canonical": raw_header_by_canonical,
        "available_fields": available_fields,
        "snapshot_label": snapshot_label,
    }


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------
def editable_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
    return str(value)


def metric_currency(value: Any) -> str:
    if value is None or value == "":
        return "-"
    try:
        return f"${float(value):,.0f}"
    except Exception:
        return str(value)


def metric_pct(value: Any) -> str:
    if value is None or value == "":
        return "-"
    try:
        return f"{float(value):.1%}"
    except Exception:
        return str(value)


def metric_date(value: Any) -> str:
    if value is None or value == "":
        return "-"
    try:
        if pd.isna(value):
            return "-"
    except Exception:
        pass
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return "-"
        return value.strftime("%m/%d/%Y")
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    parsed = maybe_parse_date(str(value))
    if parsed is not None:
        return parsed.strftime("%m/%d/%Y")
    return str(value)


def normalize_money_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    if text in {"-", "N/A", "NA"}:
        return text
    number = normalize_number(text)
    if number is None:
        return text
    return f"${number:,.0f}"


def format_edit_value(field: str, value: Any) -> str:
    text = editable_text(value)
    if field in NUMERIC_OR_TEXT_FIELDS:
        return normalize_money_text(text)
    return text


# ---------------------------------------------------------------------------
# Comment parsing / composition helpers
# ---------------------------------------------------------------------------
DATE_SPLIT_RE = re.compile(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}\s*[:\-]")
DATE_PREFIX_RE = re.compile(r"^\s*(\d{1,2}[./-]\d{1,2}[./-]\d{2,4})\s*[:\-]\s*(.*)$", re.S)


def maybe_parse_date(text: str) -> datetime | None:
    candidate = text.strip()
    if not candidate:
        return None
    for fmt in (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%m.%d.%Y",
        "%m/%d/%y",
        "%m-%d-%y",
        "%m.%d.%y",
        "%Y/%m/%d",
        "%b %d %Y",
        "%B %d %Y",
    ):
        try:
            return datetime.strptime(candidate, fmt)
        except ValueError:
            continue
    return None


def split_comment_history(text: str) -> list[str]:
    cleaned = (text or "").strip()
    if not cleaned:
        return []

    matches = list(DATE_SPLIT_RE.finditer(cleaned))
    if not matches:
        return [cleaned]

    starts = [0]
    for match in matches:
        if match.start() == 0:
            continue
        prefix = cleaned[max(0, match.start() - 4):match.start()]
        if "\n" in prefix or re.search(r"[.;:]\s*$", prefix):
            starts.append(match.start())

    starts = sorted(set(starts))
    if starts == [0]:
        return [cleaned]

    parts: list[str] = []
    for idx, start in enumerate(starts):
        end = starts[idx + 1] if idx + 1 < len(starts) else len(cleaned)
        segment = cleaned[start:end].strip()
        if segment:
            parts.append(segment)
    return parts or [cleaned]


def comment_parts(text: str) -> list[dict[str, Any]]:
    parts = []
    for segment in split_comment_history(text):
        match = DATE_PREFIX_RE.match(segment)
        parsed_date = maybe_parse_date(match.group(1)) if match else None
        parts.append(
            {
                "raw": segment,
                "date": parsed_date,
                "body": match.group(2).strip() if match else segment,
            }
        )
    return parts


def latest_history_comment(text: str) -> str:
    parts = comment_parts(text)
    if not parts:
        return ""
    return parts[0]["raw"]


def last_history_date(text: str) -> datetime | None:
    parts = comment_parts(text)
    dated = [part["date"] for part in parts if part["date"] is not None]
    return max(dated) if dated else None


def build_weekly_comment_text(detail: str, template: str = "") -> str:
    detail_clean = (detail or "").strip()
    template_clean = (template or "").strip()
    if template_clean and detail_clean:
        if detail_clean.lower().startswith(template_clean.lower()):
            return detail_clean
        return f"{template_clean}. {detail_clean}"
    return detail_clean or template_clean


def format_comment_entry_date(entry_date: date | datetime | None) -> str:
    if entry_date is None:
        entry_date = date.today()
    if isinstance(entry_date, datetime):
        entry_date = entry_date.date()
    return f"{entry_date.month}/{entry_date.day}/{str(entry_date.year)[-2:]}"


def compose_comment_value(history_text: str, this_week_comment: str, entry_date: date | datetime | None) -> str:
    history_raw = "" if history_text is None else str(history_text)
    history_clean = history_raw.strip()
    current_clean = (this_week_comment or "").strip()
    if not current_clean:
        return history_raw

    if DATE_PREFIX_RE.match(current_clean):
        new_entry = current_clean
    else:
        new_entry = f"{format_comment_entry_date(entry_date)} - {current_clean}"

    if not history_clean:
        return new_entry
    return f"{new_entry}\n{history_raw}"



# ---------------------------------------------------------------------------
# Editor dataframe helpers
# ---------------------------------------------------------------------------
def make_editor_df(deals_df: pd.DataFrame) -> pd.DataFrame:
    base_columns = ["_excel_row", *GRID_CONTEXT_HEADERS]
    export_columns = [field for field in EXPORTABLE_FIELDS if field in deals_df.columns]
    readonly_columns = [field for field in READONLY_CONTEXT_FIELDS if field in deals_df.columns]

    keep_columns = [col for col in base_columns + export_columns + readonly_columns if col in deals_df.columns]
    editor_df = deals_df[keep_columns].copy()

    for header in export_columns + readonly_columns:
        editor_df[header] = editor_df[header].apply(lambda value, header=header: format_edit_value(header, value))

    if "Comments" not in editor_df.columns:
        editor_df["Comments"] = ""
    if COMMENT_STORAGE_FIELD not in editor_df.columns:
        editor_df[COMMENT_STORAGE_FIELD] = ""

    editor_df["Comment History"] = editor_df[COMMENT_STORAGE_FIELD].fillna("").astype(str)
    editor_df["Previous Weekly Comment"] = editor_df["Comment History"].apply(latest_history_comment)
    editor_df["This Week Comment"] = ""
    editor_df["Comments Preview"] = editor_df["Comment History"]
    editor_df["Last Comment Date"] = editor_df["Comment History"].apply(last_history_date)
    editor_df["Needs Discussion"] = False

    return editor_df


def normalize_for_compare(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def refresh_editor_derived_fields(editor_df: pd.DataFrame, entry_date: date | datetime | None) -> pd.DataFrame:
    refreshed = editor_df.copy()
    for field in NUMERIC_OR_TEXT_FIELDS:
        if field in refreshed.columns:
            refreshed[field] = refreshed[field].apply(normalize_money_text)
    if COMMENT_STORAGE_FIELD in refreshed.columns:
        refreshed["Comment History"] = refreshed[COMMENT_STORAGE_FIELD].fillna("").astype(str)
    refreshed["Previous Weekly Comment"] = refreshed["Comment History"].fillna("").astype(str).apply(latest_history_comment)
    refreshed["Comments Preview"] = refreshed.apply(
        lambda row: compose_comment_value(
            row.get("Comment History", ""),
            row.get("This Week Comment", ""),
            entry_date,
        ),
        axis=1,
    )
    refreshed["Last Comment Date"] = refreshed["Comment History"].fillna("").astype(str).apply(last_history_date)
    return refreshed


def apply_detail_edit(editor_df: pd.DataFrame, excel_row: int, values: dict[str, Any], entry_date: date | datetime | None) -> pd.DataFrame:
    updated = editor_df.copy()
    mask = updated["_excel_row"] == excel_row
    for header, value in values.items():
        if header not in updated.columns:
            updated[header] = ""
        updated.loc[mask, header] = value
    return refresh_editor_derived_fields(updated, entry_date)


def unique_nonblank(series: pd.Series) -> list[str]:
    values = []
    for item in series.dropna().astype(str):
        cleaned = item.strip()
        if cleaned and cleaned not in values:
            values.append(cleaned)
    return values


def quarter_options(start_year: int, end_year: int) -> list[str]:
    values = []
    for year in range(start_year, end_year + 1):
        yy = str(year)[-2:]
        for q in [1, 2, 3, 4]:
            values.append(f"{q}Q{yy}")
    return values


def merge_options(preferred: list[str], existing: list[str]) -> list[str]:
    merged = []
    for item in preferred + existing:
        cleaned = str(item).strip()
        if cleaned and cleaned not in merged:
            merged.append(cleaned)
    return merged


def build_picklists(deals_df: pd.DataFrame) -> dict[str, list[str]]:
    today_year = date.today().year
    quarter_list = quarter_options(today_year - 1, today_year + 2)

    preferred = {
        "Resolution Likelihood": ["No", "Low", "Medium", "High", "Likely", "Resolved", "Yes"],
        "Estimated Resolution Timing": quarter_list + ["TBD", "N/A", "NEW", "NA"],
        "Estimated Liquidation Timing": quarter_list + ["TBD", "N/A", "NEW", "NA"],
        "Resolution Timing": quarter_list + ["TBD", "N/A", "NEW", "NA"],
        "Liquidity Event Timing": quarter_list + ["TBD", "N/A", "NA", "2026"],
        "Mod Needed Prior to Final Resolution (Y/N)": ["Y", "N", "N/A"],
        "FCL Needed Prior to Final Resolution (Y/N)": ["Y", "N", "Foreclosure", "Payoff", "0", "N/A"],
        "Pref Deal (Y/N)": ["Y", "N"],
        "CV Inspected in Last 6 months? (Y/N)": ["Y", "N", "N/A"],
        "New Valuation in 1Q26 (Y/N)": ["Y", "N", "N/A"],
        "Is Appraisal Finalized? (Y/N)": ["Y", "N", "N/A", "-"],
        "Third Party Offer Note": ["Yes", "No", "Y", "N", "N/A", "NA"],
        "Valuation Type": [
            "Appraisal",
            "Exterior Appraisal",
            "Interior Appraisal",
            "Drive-By BPO",
            "Exterior BPO",
            "Interior BPO",
            "Broker Opinion of Value (BOV)",
            "Desktop Valuation",
            "AVM",
            "Updated Appraisal",
            "N/A",
        ],
        "Expected Final Resolution": [
            "Payoff",
            "Payoff with Pref",
            "DPO",
            "Foreclosure",
            "Receivership Sale",
            "REO Sale",
            "Note Sale",
            "Mod/FC",
            "Mod/Paydowns",
            "JV with new developer",
            "Title Insurance",
        ],
        "Expected Resolution Type": [
            "No",
            "Resolved",
            "FC",
            "FCL",
            "FB",
            "BK",
            "REO",
            "REO Sale",
            "Settlement & DIL",
            "DIL/Refi",
            "Refi with Pref/Mezz",
            "Note Sale | Property Sale",
        ],
    }

    picklists: dict[str, list[str]] = {}
    for field, preferred_options in preferred.items():
        if field in deals_df.columns:
            picklists[field] = merge_options(preferred_options, unique_nonblank(deals_df[field]))
    picklists["Comment Template"] = COMMENT_TEMPLATE_OPTIONS
    return picklists


def nonstandard_picklist_flags(row: pd.Series, picklists: dict[str, list[str]]) -> list[str]:
    flags: list[str] = []
    for field in [
        "Mod Needed Prior to Final Resolution (Y/N)",
        "FCL Needed Prior to Final Resolution (Y/N)",
        "CV Inspected in Last 6 months? (Y/N)",
        "New Valuation in 1Q26 (Y/N)",
        "Is Appraisal Finalized? (Y/N)",
        "Valuation Type",
    ]:
        value = str(row.get(field, "")).strip()
        if value and field in picklists and value not in picklists[field]:
            flags.append(f"Nonstandard {field}")
    return flags


def row_quality_flags(row: pd.Series, picklists: dict[str, list[str]]) -> list[str]:
    flags: list[str] = []
    if not str(row.get("Resolution Likelihood", "")).strip():
        flags.append("Missing likelihood")
    if not str(row.get("Expected Final Resolution", "")).strip() and not str(row.get("Expected Resolution Type", "")).strip():
        flags.append("Missing resolution path")
    if str(row.get("Pref Deal (Y/N)", "")).strip().upper() == "Y" and not str(row.get("Pref Amount ($)", "")).strip():
        flags.append("Pref amount missing")
    if str(row.get("New Valuation in 1Q26 (Y/N)", "")).strip().upper() == "Y" and not str(row.get("Date of Valuation", "")).strip():
        flags.append("Valuation date missing")
    if not str(row.get("This Week Comment", "")).strip():
        flags.append("No current-week comment")
    flags.extend(nonstandard_picklist_flags(row, picklists))
    return flags


def build_display_view(
    base_deals: pd.DataFrame,
    editor_df: pd.DataFrame,
    property_summary: pd.DataFrame,
    entry_date: date | datetime | None,
    picklists: dict[str, list[str]],
) -> pd.DataFrame:
    display_df = base_deals.copy()
    editor_refreshed = refresh_editor_derived_fields(editor_df, entry_date)
    editor_lookup = editor_refreshed.set_index("_excel_row")

    merge_cols = [
        col
        for col in [
            *EXPORTABLE_FIELDS,
            *READONLY_CONTEXT_FIELDS,
            "Comment History",
            "Previous Weekly Comment",
            "This Week Comment",
            "Comments Preview",
            "Last Comment Date",
            "Needs Discussion",
        ]
        if col in editor_lookup.columns
    ]
    for header in merge_cols:
        display_df[header] = display_df["_excel_row"].map(editor_lookup[header].to_dict())

    if not property_summary.empty:
        display_df = display_df.merge(property_summary, on="Deal Number", how="left")

    display_df["Flag Details"] = display_df.apply(lambda row: row_quality_flags(row, picklists), axis=1)
    display_df["Flag Count"] = display_df["Flag Details"].apply(len)
    display_df["Flag Summary"] = display_df["Flag Details"].apply(lambda items: "; ".join(items[:3]) + (" ..." if len(items) > 3 else ""))

    return display_df


def workbook_change_set(
    editor_df: pd.DataFrame,
    original_editor_values: dict[int, dict[str, str]],
    entry_date: date | datetime | None,
) -> dict[tuple[int, str], str]:
    changes: dict[tuple[int, str], str] = {}
    row_lookup = refresh_editor_derived_fields(editor_df, entry_date).set_index("_excel_row")
    for excel_row, original in original_editor_values.items():
        if excel_row not in row_lookup.index:
            continue
        current = row_lookup.loc[excel_row]
        for header in EXPORTABLE_FIELDS:
            if header == COMMENT_STORAGE_FIELD:
                current_text = compose_comment_value(
                    original.get(COMMENT_STORAGE_FIELD, ""),
                    current.get("This Week Comment", ""),
                    entry_date,
                )
            elif header == "Comments":
                current_text = normalize_for_compare(original.get("Comments", ""))
            else:
                if header not in current.index:
                    continue
                current_text = normalize_for_compare(current.get(header, ""))
            original_text = normalize_for_compare(original.get(header, ""))
            if current_text != original_text:
                changes[(excel_row, header)] = current_text
    return changes


# ---------------------------------------------------------------------------
# XML patching / workbook export
# ---------------------------------------------------------------------------
def maybe_parse_number(text: str) -> int | float | None:
    candidate = text.strip()
    if not candidate:
        return None
    numeric_pattern = r"^\(?\$?[-+]?\d[\d,]*(\.\d+)?\)?$"
    if not re.match(numeric_pattern, candidate):
        return None
    negative = candidate.startswith("(") and candidate.endswith(")")
    cleaned = candidate.replace("$", "").replace(",", "").replace("(", "").replace(")", "")
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if negative:
        number *= -1
    if number.is_integer():
        return int(number)
    return number


def coerce_cell_value(text: str, header: str) -> Any:
    if text is None:
        return None
    if isinstance(text, float) and math.isnan(text):
        return None
    if not isinstance(text, str):
        return text
    if text == "":
        return None

    if header in DATE_OR_TEXT_FIELDS:
        parsed_date = maybe_parse_date(text)
        if parsed_date is not None:
            return parsed_date
        return text

    if header in NUMERIC_OR_TEXT_FIELDS:
        parsed_number = maybe_parse_number(text)
        if parsed_number is not None:
            return parsed_number
        return text

    return text


def qn(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def get_or_create_cell(row_elem: ET.Element, coord: str) -> ET.Element:
    target_idx = column_index_from_string(re.sub(r"\d", "", coord))
    cell_tag = qn(MAIN_NS, "c")

    for cell in row_elem.findall(cell_tag):
        if cell.get("r") == coord:
            return cell

    new_cell = ET.Element(cell_tag, {"r": coord})
    inserted = False
    for pos, cell in enumerate(list(row_elem)):
        existing_ref = cell.get("r")
        if not existing_ref:
            continue
        existing_idx = column_index_from_string(re.sub(r"\d", "", existing_ref))
        if existing_idx > target_idx:
            row_elem.insert(pos, new_cell)
            inserted = True
            break
    if not inserted:
        row_elem.append(new_cell)
    return new_cell


def set_cell_xml_value(cell_elem: ET.Element, value: Any) -> None:
    for child in list(cell_elem):
        cell_elem.remove(child)
    cell_elem.attrib.pop("t", None)

    if value is None:
        return

    if isinstance(value, datetime):
        v = ET.SubElement(cell_elem, qn(MAIN_NS, "v"))
        v.text = str(to_excel(value))
        return

    if isinstance(value, date):
        v = ET.SubElement(cell_elem, qn(MAIN_NS, "v"))
        v.text = str(to_excel(datetime.combine(value, datetime.min.time())))
        return

    if isinstance(value, bool):
        cell_elem.set("t", "b")
        v = ET.SubElement(cell_elem, qn(MAIN_NS, "v"))
        v.text = "1" if value else "0"
        return

    if isinstance(value, (int, float)):
        v = ET.SubElement(cell_elem, qn(MAIN_NS, "v"))
        v.text = str(value)
        return

    cell_elem.set("t", "inlineStr")
    is_elem = ET.SubElement(cell_elem, qn(MAIN_NS, "is"))
    t_elem = ET.SubElement(is_elem, qn(MAIN_NS, "t"))
    t_elem.set(f"{{{XML_SPACE}}}space", "preserve")
    t_elem.text = str(value)


def patch_sheet_xml(sheet_xml: bytes, changes: dict[tuple[int, str], str], cell_meta: dict[tuple[int, str], dict[str, Any]]) -> bytes:
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(qn(MAIN_NS, "sheetData"))
    if sheet_data is None:
        return sheet_xml

    row_map = {int(row.get("r")): row for row in sheet_data.findall(qn(MAIN_NS, "row"))}

    for (excel_row, header), text_value in changes.items():
        meta = cell_meta.get((excel_row, header))
        if not meta:
            continue
        row_elem = row_map.get(excel_row)
        if row_elem is None:
            row_elem = ET.SubElement(sheet_data, qn(MAIN_NS, "row"), {"r": str(excel_row)})
            row_map[excel_row] = row_elem
        cell_elem = get_or_create_cell(row_elem, meta["coord"])
        coerced = coerce_cell_value(text_value, header)
        set_cell_xml_value(cell_elem, coerced)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def patch_workbook_xml(workbook_xml: bytes) -> bytes:
    root = ET.fromstring(workbook_xml)
    calc_pr = root.find(qn(MAIN_NS, "calcPr"))
    if calc_pr is None:
        calc_pr = ET.SubElement(root, qn(MAIN_NS, "calcPr"))
    calc_pr.set("calcMode", "auto")
    calc_pr.set("calcOnSave", "1")
    calc_pr.attrib.pop("fullCalcOnLoad", None)
    calc_pr.attrib.pop("forceFullCalc", None)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def patch_workbook_rels_xml(rels_xml: bytes) -> bytes:
    root = ET.fromstring(rels_xml)
    calcchain_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain"
    for rel in list(root):
        if rel.get("Type") == calcchain_type:
            root.remove(rel)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def patch_content_types_xml(content_xml: bytes) -> bytes:
    root = ET.fromstring(content_xml)
    calc_part = "/xl/calcChain.xml"
    for child in list(root):
        if child.get("PartName") == calc_part:
            root.remove(child)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)



def build_updated_workbook(
    file_bytes: bytes,
    sheet_path: str,
    changes: dict[tuple[int, str], str],
    cell_meta: dict[tuple[int, str], dict[str, Any]],
    sheet_name: str | None = None,
) -> bytes:
    """
    Safe workbook writer that round-trips through openpyxl instead of patching raw XML.
    This is more tolerant of SharePoint/Excel-generated workbooks and avoids corruption on download.
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    target_ws = None

    if sheet_name:
        for ws in wb.worksheets:
            if ws.title == sheet_name:
                target_ws = ws
                break

    # Fall back to the first worksheet that contains known coordinates if the XML sheet path
    # does not map cleanly back to a visible worksheet title.
    if target_ws is None and cell_meta:
        coord_sample = next(iter(cell_meta.values())).get("coord")
        for ws in wb.worksheets:
            try:
                _ = ws[coord_sample]
                target_ws = ws
                break
            except Exception:
                continue

    if target_ws is None:
        raise ValueError(f"Could not map worksheet path {sheet_path} back to a workbook sheet.")

    for (excel_row, header), text_value in changes.items():
        meta = cell_meta.get((excel_row, header))
        if not meta:
            continue
        coord = meta.get("coord")
        if not coord:
            continue
        target_ws[coord].value = coerce_cell_value(text_value, header)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()



# ---------------------------------------------------------------------------
# Misc display helpers
# ---------------------------------------------------------------------------
def format_deal_label(row: pd.Series) -> str:
    am = editable_text(row.get("Asset Manager")) or "No AM"
    dq = editable_text(row.get("Current DQ Status")) or "No DQ"
    return f"{row['Deal Name']} | {am} | {dq} | #{row['Deal Number']}"


def sort_deals(df: pd.DataFrame, mode: str) -> pd.DataFrame:
    working = df.copy()
    if mode == "UPB desc":
        working["_sort_upb"] = working["Current UPB"].apply(normalize_number)
        return working.sort_values(["_sort_upb", "Asset Manager", "Deal Name"], ascending=[False, True, True]).drop(columns=["_sort_upb"])
    if mode == "Flag count desc":
        return working.sort_values(["Flag Count", "Asset Manager", "Deal Name"], ascending=[False, True, True])
    if mode == "Last comment date desc":
        return working.sort_values(["Last Comment Date", "Asset Manager", "Deal Name"], ascending=[False, True, True])
    if mode == "Deal Name":
        return working.sort_values(["Deal Name", "Asset Manager", "Deal Number"])
    return working.sort_values(["Asset Manager", "Deal Name", "Deal Number"])

